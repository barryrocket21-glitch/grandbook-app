#!/usr/bin/env python3
"""
One-time import: SPX shipping report Mei 2026 → orders + order_items.
Source: ~/Downloads/d93ff4578a6242dcbd457a80321ed113.xlsx (1554 row, May 1-23).

Strategy:
- Bulk REST insert via Supabase service role.
- Triggers compute_commissions + compute_order_costs fire AFTER INSERT
  → auto-generate commission + cost estimates.
- Phone normalize: strip 62 prefix, ensure leading 0.
- Status map: SPX raw → GrandBook internal.
- Product matcher: prefix-based.
- order_number generated per-day-counter (GB-YYYYMMDD-NNNNNN).
"""
import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import urllib.request
import urllib.error

# ---------- Config ----------
FILE = Path.home() / 'Downloads/d93ff4578a6242dcbd457a80321ed113.xlsx'
ENV_FILE = Path('/Users/mr.nobody/GrandBook/grandbook-app/.env.local')
ORG_ID = 1
CHANNEL_ID = 1  # SPX_DIRECT
LISA_ID = '42073cb0-6406-4edb-9a27-56a445b158d9'
BARRY_ID = 'f8726f49-5d7b-4029-9ad0-2f414dd96717'
BATCH_SIZE = 100

PRODUCT_MAP = {
    'Nature Gemuk Badan': 7,
    'Kran Robotic Arm': 6,
    'Kran Robotic': 6,  # variant suffix like G
    'Jaring Paranet': 5,
    'TM Jaring Paranet': 5,
    'Shampo Uban': 9,
    'Sulem': 10,
    'MJO Luna': 8,
    'Sandal Luna': 8,
    'Blade K45': 11,  # Newly created
}

# Explicitly skip these (Barry decide later via /products UI)
PRODUCT_SKIP = ['Kran Waterfall']

STATUS_MAP = {
    'Delivered': 'DITERIMA',
    'In Transit': 'DIKIRIM',
    'Returned': 'RETUR',
    'Returning': 'DIKIRIM',  # in process of return
    'Cancelled': 'CANCEL',
    'Pending Pickup': 'SIAP_KIRIM',
    'Pickup On Hold': 'SIAP_KIRIM',
    'On Hold': 'PROBLEM',
    'Delivering': 'DIKIRIM',
}

RESI_STATUS_MAP = {
    'DITERIMA': 'DITERIMA',
    'DIKIRIM': 'AKTIF',
    'RETUR': 'RETUR',
    'CANCEL': 'AKTIF',  # canceled but resi exists
    'PROBLEM': 'PROBLEM',
    'SIAP_KIRIM': 'AKTIF',
}

# ---------- Load env ----------
env = {}
for line in ENV_FILE.read_text().splitlines():
    if '=' in line and not line.startswith('#'):
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip()

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SERVICE_KEY = env['SUPABASE_SERVICE_ROLE_KEY']


# ---------- Helpers ----------
def normalize_phone(raw):
    if not raw:
        return None
    s = str(raw).strip()
    # Strip any non-digit
    s = ''.join(c for c in s if c.isdigit())
    if not s:
        return None
    # Strip 62 prefix
    if s.startswith('62') and len(s) > 10:
        s = '0' + s[2:]
    elif s.startswith('8') and len(s) >= 9:
        s = '0' + s
    elif not s.startswith('0'):
        s = '0' + s
    return s


def parse_dt(raw):
    """Parse '23-05-2026 13:22' → ISO timestamp."""
    if not raw or raw == '-':
        return None
    s = str(raw).strip()
    try:
        dt = datetime.strptime(s, '%d-%m-%Y %H:%M')
        return dt.isoformat()
    except ValueError:
        try:
            dt = datetime.strptime(s, '%Y-%m-%d %H:%M')
            return dt.isoformat()
        except ValueError:
            try:
                dt = datetime.strptime(s, '%Y-%m-%d')
                return dt.isoformat()
            except ValueError:
                return None


def parse_date_only(raw):
    """Parse '23-05-2026 13:22' or '23-05-2026' → 'YYYY-MM-DD'."""
    if not raw or raw == '-':
        return None
    s = str(raw).strip().split(' ')[0]
    try:
        dt = datetime.strptime(s, '%d-%m-%Y')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        return None


def match_product(item_str):
    if not item_str:
        return None
    s = str(item_str).strip()
    # Strip qty prefix "Nx " (e.g. "1x MJO Luna")
    import re
    s = re.sub(r'^\d+x\s+', '', s)
    # Strip brand prefix "MB " / "TM "
    s = re.sub(r'^(MB|TM)\s+', '', s)
    # Explicit skip list
    for skip in PRODUCT_SKIP:
        if skip in s:
            return None
    # Substring match — find longest matching key
    for key in sorted(PRODUCT_MAP.keys(), key=len, reverse=True):
        if key in s:
            return PRODUCT_MAP[key]
    return None


def to_float(v, default=0):
    if v is None or v == '' or v == '-':
        return default
    try:
        return float(str(v).replace(',', ''))
    except (ValueError, TypeError):
        return default


def to_int(v, default=1):
    try:
        return int(float(str(v).replace(',', '')))
    except (ValueError, TypeError):
        return default


def post_rest(table, payload):
    """POST to PostgREST endpoint, returns parsed JSON list of inserted rows."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    body = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(
        url,
        data=body,
        headers={
            'apikey': SERVICE_KEY,
            'Authorization': f'Bearer {SERVICE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'return=representation',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8')
        print(f"ERROR {e.code}: {err_body[:500]}", file=sys.stderr)
        raise


# ---------- Parse XLSX ----------
print(f"Loading {FILE}...")
wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb['Sheet1']

rows = []
unmapped_products = defaultdict(int)
unmapped_statuses = defaultdict(int)

for r_idx in range(3, ws.max_row + 1):  # skip row 1 (download time) + row 2 (header)
    resi = ws.cell(row=r_idx, column=1).value
    if not resi or not str(resi).startswith('SPXID'):
        continue

    raw_status = ws.cell(row=r_idx, column=6).value
    create_time = ws.cell(row=r_idx, column=5).value  # DD-MM-YYYY HH:MM
    pickup_time = ws.cell(row=r_idx, column=11).value
    delivered_time = ws.cell(row=r_idx, column=12).value
    returning_time = ws.cell(row=r_idx, column=15).value

    recipient_name = ws.cell(row=r_idx, column=16).value
    recipient_phone = ws.cell(row=r_idx, column=17).value
    recipient_prov = ws.cell(row=r_idx, column=18).value
    recipient_city = ws.cell(row=r_idx, column=19).value
    recipient_district = ws.cell(row=r_idx, column=20).value
    recipient_address = ws.cell(row=r_idx, column=21).value
    recipient_zip = ws.cell(row=r_idx, column=22).value

    item_in_parcel = ws.cell(row=r_idx, column=32).value  # e.g. "Jaring Paranet Ukuran: ..."
    qty = ws.cell(row=r_idx, column=33).value
    cod_collection = ws.cell(row=r_idx, column=34).value  # Y/N
    cod_amount = ws.cell(row=r_idx, column=35).value
    parcel_value = ws.cell(row=r_idx, column=36).value
    parcel_weight = ws.cell(row=r_idx, column=37).value
    actual_weight = ws.cell(row=r_idx, column=38).value
    est_ship = ws.cell(row=r_idx, column=39).value
    actual_ship = ws.cell(row=r_idx, column=40).value
    cod_service_fee = ws.cell(row=r_idx, column=43).value
    return_ship = ws.cell(row=r_idx, column=44).value

    # Status mapping
    internal_status = STATUS_MAP.get(str(raw_status))
    if not internal_status:
        unmapped_statuses[str(raw_status)] += 1
        continue

    # Product mapping
    product_id = match_product(item_in_parcel)
    if not product_id:
        unmapped_products[str(item_in_parcel)[:60]] += 1
        continue

    # Phone
    phone = normalize_phone(recipient_phone)

    # Build payload
    order_date = parse_date_only(create_time)
    if not order_date:
        continue

    payout_amount = None
    cod_settled_at = None
    if internal_status == 'DITERIMA':
        # Payout = COD Amount - actual_shipping_fee - cod_service_fee
        payout_amount = to_float(cod_amount) - to_float(actual_ship) - to_float(cod_service_fee)

    rows.append({
        '_row_idx': r_idx,
        '_resi': str(resi),
        '_product_id': product_id,
        '_qty': to_int(qty, 1),
        '_item_in_parcel': str(item_in_parcel),
        'order': {
            'organization_id': ORG_ID,
            'channel_id': CHANNEL_ID,
            'cs_id': LISA_ID,
            'created_by': BARRY_ID,
            'external_order_id': str(resi),  # Use SPX tracking as external ID for dedupe
            'status': internal_status,
            'resi': str(resi),
            'cs_name': 'lisa',
            'customer_name': str(recipient_name) if recipient_name else '',
            'customer_phone': phone,
            'customer_province': str(recipient_prov) if recipient_prov else None,
            'customer_city': str(recipient_city) if recipient_city else None,
            'customer_subdistrict': str(recipient_district) if recipient_district else None,
            'customer_address_detail': str(recipient_address) if recipient_address else None,
            'customer_zip': str(int(recipient_zip)) if recipient_zip and str(recipient_zip).replace('.','').isdigit() else (str(recipient_zip) if recipient_zip else None),
            'payment_method': 'COD' if str(cod_collection).upper() == 'Y' else 'TRANSFER',
            'cod_amount': to_float(cod_amount) if str(cod_collection).upper() == 'Y' else None,
            'total': to_float(parcel_value),
            'shipping_cost': to_float(est_ship),
            'shipping_cost_actual': to_float(actual_ship) if internal_status in ('DITERIMA','RETUR') else None,
            'payout_amount': payout_amount,
            'order_date': order_date,
            'resi_printed_at': parse_dt(create_time),
            'picked_up_at': parse_dt(pickup_time),
            'delivered_at': parse_dt(delivered_time) if internal_status == 'DITERIMA' else None,
            'returned_at': parse_dt(returning_time) if internal_status == 'RETUR' else None,
            'status_changed_at': (
                parse_dt(delivered_time) if internal_status == 'DITERIMA'
                else parse_dt(returning_time) if internal_status == 'RETUR'
                else parse_dt(pickup_time) if internal_status == 'DIKIRIM'
                else parse_dt(create_time)
            ),
        },
        '_item': {
            'organization_id': ORG_ID,
            'product_id': product_id,
            'qty': to_int(qty, 1),
            'price': to_float(parcel_value),
            'product_name_raw': str(item_in_parcel),
            'weight_per_unit': to_float(parcel_weight) if parcel_weight else None,
        },
    })

print(f"\nParsed {len(rows)} valid rows")
if unmapped_statuses:
    print(f"\nUnmapped statuses ({sum(unmapped_statuses.values())} rows skipped):")
    for s, c in unmapped_statuses.items():
        print(f"  {s}: {c}")
if unmapped_products:
    print(f"\nUnmapped products ({sum(unmapped_products.values())} rows skipped):")
    for p, c in unmapped_products.items():
        print(f"  {p}: {c}")


# ---------- Generate order_number per day counter ----------
day_counter = defaultdict(int)
for r in rows:
    od = r['order']['order_date']  # YYYY-MM-DD
    day_counter[od] += 1
    counter = day_counter[od]
    yyyymmdd = od.replace('-', '')
    r['order']['order_number'] = f"GB-{yyyymmdd}-{counter:06d}"


# ---------- Insert in batches ----------
print(f"\n--- Inserting orders (batch size {BATCH_SIZE}) ---")
inserted_count = 0
all_items = []

for i in range(0, len(rows), BATCH_SIZE):
    batch = rows[i:i+BATCH_SIZE]
    payload = [r['order'] for r in batch]
    try:
        result = post_rest('orders', payload)
        # Map inserted IDs back to items
        for r, inserted in zip(batch, result):
            all_items.append({
                'order_id': inserted['id'],
                **r['_item'],
            })
        inserted_count += len(result)
        print(f"  [{i+len(batch):4d}/{len(rows)}] inserted")
    except Exception as e:
        print(f"  ERROR at batch {i}: {e}", file=sys.stderr)
        # Skip on error, continue
        continue

print(f"\nTotal orders inserted: {inserted_count}")

# ---------- Insert order_items in batches ----------
print(f"\n--- Inserting order_items ---")
item_inserted = 0
for i in range(0, len(all_items), BATCH_SIZE):
    batch = all_items[i:i+BATCH_SIZE]
    try:
        result = post_rest('order_items', batch)
        item_inserted += len(result)
        print(f"  [{i+len(batch):4d}/{len(all_items)}] items inserted")
    except Exception as e:
        print(f"  ERROR at items batch {i}: {e}", file=sys.stderr)

print(f"\n=== DONE ===")
print(f"Orders inserted: {inserted_count}")
print(f"Items inserted: {item_inserted}")
